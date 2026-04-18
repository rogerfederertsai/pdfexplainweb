# -*- coding: utf-8 -*-
"""
NEW 版 Excel 輸出服務：
- 解析結果**僅寫入** INDEX.xlsx（土地→「土地」工作表、建物→「建物」工作表）；不另產個案 .xlsx。
- SAMPLE.xlsx 僅供表頭對照；輸出資料夾內可保留 SAMPLE.xlsx 與 INDEX.xlsx。
- INDEX 儲存格字型使用標楷體；Web 存檔時可寫入使用者預填之「備註」（CLI 自動寫入仍為空白）。
- 「土地」「建物」工作表：凍結表頭列、表頭鎖定並啟用無密碼工作表保護（資料列可編輯；Excel 可「取消保護工作表」後改表頭）。

完全獨立於任何前端 Web 或 Supabase，僅專注於檔案輸出格式。
"""

from __future__ import annotations

import logging
import re
import threading
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from config import INDEX_XLSX, OUTPUT_BY_SECTION_LOT, SAMPLE_XLSX
from services.parser import safe_filename

logger = logging.getLogger(__name__)

# INDEX 與輸出儲存格字型（Windows 標楷體）
try:
    from openpyxl.styles import Font

    _FONT_INDEX = Font(name="標楷體", size=11)
except Exception:  # openpyxl 未載入時略過
    _FONT_INDEX = None

# INDEX 讀寫鎖：避免多執行緒同時改寫編號（Pipeline 仍以主執行緒寫入為主，此為保險）
_INDEX_LOCK = threading.Lock()

# 舊版 INDEX 僅有單一工作表「INDEX」且表頭在第 1 列（遷移時辨識用）
_LEGACY_INDEX_SHEET = "INDEX"


def _resolve_sample_path() -> Path:
    """取得 SAMPLE.xlsx 路徑；若常數路徑不存在則於專案內 rglob 搜尋（避開資料夾編碼問題）。"""
    if SAMPLE_XLSX.exists():
        return SAMPLE_XLSX
    root = OUTPUT_BY_SECTION_LOT.parent.resolve()
    found = next(root.rglob("SAMPLE.xlsx"), None)
    if found and found.is_file():
        return found
    return SAMPLE_XLSX


def _parse_area_sqm_number(area_str: Any) -> Optional[float]:
    """從「2,670 平方公尺」或純數字取出平方公尺數值，供 SAMPLE 之數字欄使用。"""
    if area_str is None:
        return None
    s = str(area_str).strip()
    if not s:
        return None
    m = re.search(r"([\d,.]+)\s*平方公尺", s)
    if m:
        s = m.group(1)
    s = s.replace(",", "")
    try:
        v = float(s)
        return float(int(v)) if v == int(v) else v
    except ValueError:
        return None


def _detect_header_row(ws) -> int:
    """偵測橫向表頭所在列（SAMPLE 為第 1 列；舊版可能為第 3 列）。"""
    v1 = ws.cell(row=1, column=1).value
    v3 = ws.cell(row=3, column=1).value
    if v1 == "編號":
        return 1
    if v3 == "編號":
        return 3
    return 1


def _apply_index_sheet_header_freeze_and_protection(ws) -> None:
    """
    凍結表頭列下方捲動區、鎖定表頭列；資料區（含表頭下若干空白列）解鎖並啟用無密碼保護。
    使用者仍可於 Excel「校閱 → 取消保護工作表」後編輯表頭。
    """
    try:
        from openpyxl.styles import Protection
        from openpyxl.worksheet.protection import SheetProtection
    except ImportError:
        return

    hr = _detect_header_row(ws)
    # 重設保護狀態（範本若帶密碼雜湊，僅設 sheet=False 可能無法解除寫入限制）
    ws.protection = SheetProtection()

    # 右側「備註」等欄須納入 locked 範圍，欄數採較寬上限
    mc = min(max(ws.max_column or 0, 48), 256)
    mr = max(ws.max_row or 0, hr)
    # 保留表頭下既有資料列，並多解鎖數列供手動新增（避免保護後無法在空白列輸入）
    mr_unlock = max(mr, hr + 1) + 150

    for r in range(hr + 1, mr_unlock + 1):
        for c in range(1, mc + 1):
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    for c in range(1, mc + 1):
        ws.cell(row=hr, column=c).protection = Protection(locked=True)

    _safe_set_freeze_below_header(ws, hr)

    ws.protection = SheetProtection(
        sheet=True,
        password=None,
        selectLockedCells=True,
        selectUnlockedCells=True,
        formatCells=False,
        formatColumns=False,
        formatRows=False,
        insertColumns=False,
        insertRows=False,
        insertHyperlinks=False,
        deleteColumns=False,
        deleteRows=False,
        sort=False,
        autoFilter=False,
        pivotTables=False,
    )


def _apply_index_workbook_header_freeze_and_protection(wb) -> None:
    """對 INDEX 工作簿內「土地」「建物」套用表頭凍結與保護。"""
    for name in ("土地", "建物"):
        if name in wb.sheetnames:
            _apply_index_sheet_header_freeze_and_protection(wb[name])


def _apply_kai_font_to_used_range(ws, min_row: int = 1, max_row: int = 500) -> None:
    """將工作表已使用範圍字型設為標楷體（供 INDEX 使用）。"""
    if _FONT_INDEX is None:
        return
    mc = ws.max_column or 15
    for r in range(min_row, max_row + 1):
        for c in range(1, mc + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None and str(cell.value).strip() != "":
                cell.font = _FONT_INDEX


def _truncate_rows_after_header(ws, header_last_row: int | None = None) -> None:
    """刪除表頭列以下所有列（建立 INDEX 自 SAMPLE 複製時，徹底移除示範資料列）。"""
    hr = header_last_row if header_last_row is not None else _detect_header_row(ws)
    max_r = ws.max_row or 0
    if max_r <= hr:
        return
    ws.delete_rows(hr + 1, max_r - hr)


def _ensure_building_sheet_with_headers(wb) -> None:
    """確保存在「建物」工作表；若自 SAMPLE 載入則保留第 1 列表頭並清除其下資料列。"""
    if "建物" in wb.sheetnames:
        ws = wb["建物"]
        hr = _detect_header_row(ws)
        _truncate_rows_after_header(ws, header_last_row=hr)
        return
    ins = 1
    if "土地" in wb.sheetnames:
        ins = wb.sheetnames.index("土地") + 1
    wb.create_sheet("建物", ins)


def _header_flat(header: Any) -> str:
    """表頭單行化（略過換行）供比對。"""
    if header is None:
        return ""
    return str(header).replace("\n", "").replace("\r", "").strip()


def _is_remark_header(header: Any) -> bool:
    """辨識「備註」欄表頭（含全形空白、NFKC 正規化）。"""
    if header is None:
        return False
    raw = unicodedata.normalize("NFKC", str(header))
    hn = raw.replace("\n", "").replace("\r", "").strip()
    if hn == "備註":
        return True
    compact = hn.replace(" ", "").replace("\u3000", "")
    return compact == "備註"


def _find_remark_column(ws, header_row: int) -> Optional[int]:
    """自表頭列掃描備註欄欄號（靠右末欄亦須掃到）。"""
    # 與 _write_ctx_to_row 一致，避免 max_column 過小漏掉僅有表頭占用的右側欄
    max_c = min(max(ws.max_column or 0, 1) + 80, 256)
    for c in range(1, max_c + 1):
        hdr = ws.cell(row=header_row, column=c).value
        if hdr is None or str(hdr).strip() == "":
            continue
        if _is_remark_header(hdr):
            return c
    return None


def _ensure_sheet_view_has_selection(ws) -> None:
    """部分由 Excel 另存的檔案 sheetView 無 selection，freeze_panes 會 IndexError；補上預設。"""
    try:
        from openpyxl.worksheet.views import Selection

        view = ws.sheet_view
        sel = getattr(view, "selection", None)
        if not sel:
            view.selection = [Selection(pane=None, activeCell="A1", sqref="A1")]
    except Exception as ex:
        logger.warning("[EXCEL] 補齊 sheetView.selection 失敗：%s", ex)


def _safe_set_freeze_below_header(ws, header_row: int) -> None:
    """凍結表頭列（凍結於表頭下一列第 1 欄）。"""
    _ensure_sheet_view_has_selection(ws)
    try:
        top_left = ws.cell(row=header_row + 1, column=1).coordinate
        ws.freeze_panes = top_left
    except Exception as ex:
        logger.warning("[EXCEL] freeze_panes 設定失敗（表頭列=%s）：%s", header_row, ex)


# 與 SAMPLE 一致、不給使用者編輯的欄位（編號由存檔時遞增；備註不寫入）
_SAMPLE_NON_EDITABLE_HEADERS = frozenset({"編號", "備註"})


def _sample_sheet_headers_in_order(ws) -> List[str]:
    """讀取工作表橫向表頭列（由左而右），保留儲存格原始字串（含換行）供畫面顯示。"""
    hr = _detect_header_row(ws)
    headers: List[str] = []
    c = 1
    while c <= 40:
        v = ws.cell(row=hr, column=c).value
        if v is None or str(v).strip() == "":
            if c > 22:
                break
            c += 1
            continue
        headers.append(str(v).strip())
        c += 1
    return headers


def _get_sample_worksheet_for_doc_kind(wb, doc_kind: str):
    """依文件類型取得 SAMPLE 中對應工作表（優先表名「土地」「建物」，否則依表頭關鍵字猜測）。"""
    dk = str(doc_kind or "land").lower()
    if dk == "building":
        if "建物" in wb.sheetnames:
            return wb["建物"]
        for name in wb.sheetnames:
            ws = wb[name]
            hs = _sample_sheet_headers_in_order(ws)
            if any("地段建號" in _header_flat(h) for h in hs):
                return ws
        return wb[wb.sheetnames[-1]] if wb.sheetnames else None
    if "土地" in wb.sheetnames:
        return wb["土地"]
    for name in wb.sheetnames:
        ws = wb[name]
        hs = _sample_sheet_headers_in_order(ws)
        flat = "".join(_header_flat(h) for h in hs)
        if "地段號" in flat and "地段建號" not in flat:
            return ws
    return wb[wb.sheetnames[0]] if wb.sheetnames else None


def sample_header_to_editable_storage(
    doc_kind: str, header_cell: Any
) -> Optional[Tuple[str, str, str]]:
    """
    將 SAMPLE 表頭對應到前端／API 使用的儲存位置。
    回傳 (bucket, dict_key, label_for_display)；bucket 為 mark | ownership | building。
    若為編號、備註或無法對應則回傳 None。
    """
    if header_cell is None:
        return None
    label = str(header_cell).strip()
    hn = _header_flat(header_cell)
    if not hn or hn in _SAMPLE_NON_EDITABLE_HEADERS:
        return None
    if label in _SAMPLE_NON_EDITABLE_HEADERS:
        return None

    dk = str(doc_kind or "land").lower()

    if dk == "building":
        if hn == "地段建號" or label == "地段建號":
            return ("building", "地段建號", label)
        if hn == "建物門牌" or label == "建物門牌":
            return ("building", "建物門牌", label)
        if "主建物面積" in hn and "平方公尺" in hn:
            return ("building", "主建物面積", label)
        if "附屬建物面積" in hn and "平方公尺" in hn:
            return ("building", "附屬建物面積", label)
        if "共有部分面積" in hn and "平方公尺" in hn:
            return ("building", "共有部分面積", label)
        if "權狀面積" in hn and "平方公尺" in hn:
            return ("building", "權狀面積", label)
        if hn == "主要用途" or label == "主要用途":
            return ("building", "主要用途", label)
    else:
        if hn == "地段號" or label == "地段號":
            return ("mark", "地段地號", label)
        if _header_matches_land_area(header_cell):
            return ("mark", "面積", label)

    if hn == "所有權人" or label == "所有權人":
        return ("ownership", "所有權人", label)
    if hn == "統一編號" or label == "統一編號":
        return ("ownership", "統一編號", label)
    if hn == "地址" or label == "地址":
        return ("ownership", "地址", label)
    if hn == "權利範圍" or label == "權利範圍":
        return ("ownership", "權利範圍", label)
    if hn == "登記原因" or label == "登記原因":
        return ("ownership", "登記原因", label)
    if hn == "登記日期" or label == "登記日期":
        return ("ownership", "登記日期", label)
    if hn == "查詢時間" or label == "查詢時間":
        return ("ownership", "查詢時間", label)
    return None


def get_sample_editable_headers_in_order(doc_kind: str) -> List[str]:
    """自 SAMPLE.xlsx 讀取指定類型工作表之表頭，排除編號與備註，順序與試算表欄位相同。"""
    import openpyxl

    path = _resolve_sample_path()
    if not path.is_file():
        return []
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = _get_sample_worksheet_for_doc_kind(wb, doc_kind)
        if ws is None:
            return []
        raw = _sample_sheet_headers_in_order(ws)
        out: List[str] = []
        for h in raw:
            if _header_flat(h) in _SAMPLE_NON_EDITABLE_HEADERS:
                continue
            if sample_header_to_editable_storage(doc_kind, h) is not None:
                out.append(h)
        return out
    finally:
        wb.close()


def _preview_value_for_cell(
    doc_kind: str,
    bucket: str,
    key: str,
    mk: Dict[str, Any],
    ow: Dict[str, Any],
    bd: Dict[str, Any],
    section: str,
    lot: str,
) -> str:
    """組合單一預覽欄位顯示字串。"""
    mk = mk or {}
    ow = ow or {}
    bd = bd or {}
    if bucket == "mark":
        if key == "地段地號":
            v = mk.get("地段地號")
            if v:
                return str(v)
            sfx = "建號" if str(doc_kind).lower() == "building" else "地號"
            return f"{section} {lot} {sfx}".strip() if (section or lot) else ""
        v = mk.get(key)
        return "" if v is None else str(v)
    if bucket == "ownership":
        v = ow.get(key)
        if key == "查詢時間" and not v:
            v = mk.get("查詢時間")
        return "" if v is None else str(v)
    if bucket == "building":
        if key == "地段建號":
            v = bd.get("地段建號")
            if v:
                return str(v)
            return f"{section} {lot} 建號".strip() if (section or lot) else ""
        v = bd.get(key)
        if v is None:
            return ""
        if isinstance(v, float) and v == int(v):
            return str(int(v))
        return str(v)
    return ""


def build_sample_aligned_preview_fields(
    doc_kind: str,
    excel_data_mark: Optional[Dict[str, Any]],
    excel_data_ownership: Optional[Dict[str, Any]],
    excel_data_building: Optional[Dict[str, Any]],
    section: str,
    lot: str,
) -> List[Dict[str, Any]]:
    """
    產生與 SAMPLE.xlsx 同欄位、同順序之預覽列，供 Web 左欄僅顯示可編輯項。
    每筆：header（顯示用，可含換行）、bucket、key、value。
    """
    headers = get_sample_editable_headers_in_order(doc_kind)
    mk = dict(excel_data_mark or {})
    ow = dict(excel_data_ownership or {})
    bd = dict(excel_data_building or {})
    rows: List[Dict[str, Any]] = []
    for h in headers:
        mapped = sample_header_to_editable_storage(doc_kind, h)
        if not mapped:
            continue
        bucket, key, _label = mapped
        val = _preview_value_for_cell(doc_kind, bucket, key, mk, ow, bd, section, lot)
        rows.append(
            {
                "header": h,
                "bucket": bucket,
                "key": key,
                "value": val,
            }
        )
    return rows


def infer_section_lot_for_save(
    doc_kind: str,
    excel_data_mark: Dict[str, Any],
    excel_data_building: Optional[Dict[str, Any]],
) -> Tuple[str, str]:
    """
    從 SAMPLE 對應之單一欄位（地段地號／地段建號）推回 地段、地號，供 INDEX 與 API 使用。
    """
    dk = str(doc_kind or "land").lower()
    mk = excel_data_mark or {}
    bd = excel_data_building or {}
    if dk == "building":
        label = str(bd.get("地段建號") or "").strip()
    else:
        label = str(mk.get("地段地號") or "").strip()
    label = re.sub(r"\s*地號\s*$", "", label)
    label = re.sub(r"\s*建號\s*$", "", label)
    label = label.strip()
    if not label:
        return "", ""
    m = re.search(r"^(.+?)\s+([\d\-]+)\s*$", label)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return "", label


def merge_preview_field_rows_to_excel(
    rows: List[Dict[str, Any]],
) -> Tuple[Dict[str, Any], Dict[str, Any], Dict[str, Any]]:
    """將前端回傳之 {header,bucket,key,value} 列還原為 excel_data_mark／ownership／building。"""
    mk: Dict[str, Any] = {}
    ow: Dict[str, Any] = {}
    bd: Dict[str, Any] = {}
    for r in rows:
        if not isinstance(r, dict):
            continue
        b = str(r.get("bucket") or "")
        k = str(r.get("key") or "")
        v = r.get("value")
        if not k:
            continue
        sval = "" if v is None else str(v)
        if b == "mark":
            mk[k] = sval
        elif b == "ownership":
            ow[k] = sval
        elif b == "building":
            bd[k] = sval
    return mk, ow, bd


def apply_sample_preview_overlays(
    base_mark: Optional[Dict[str, Any]],
    base_ownership: Optional[Dict[str, Any]],
    base_building: Optional[Dict[str, Any]],
    rows: List[Dict[str, Any]],
) -> Tuple[Dict[str, Any], Dict[str, Any], Dict[str, Any]]:
    """
    以解析快照為底，僅覆寫 SAMPLE 上可編輯之欄位（避免遺失未顯示之鍵如登記次序）。
    """
    mk = dict(base_mark or {})
    ow = dict(base_ownership or {})
    bd = dict(base_building or {})
    pm, po, pb = merge_preview_field_rows_to_excel(rows)
    mk.update(pm)
    ow.update(po)
    bd.update(pb)
    return mk, ow, bd


def _header_matches_land_area(header: Any) -> bool:
    """土地工作表「面積(平方公尺)」欄；排除建物各面積子欄。"""
    hn = _header_flat(header)
    if not hn:
        return False
    if any(
        k in hn
        for k in ("主建物面積", "附屬建物面積", "共有部分面積", "權狀面積")
    ):
        return False
    if "面積" in hn and "平方公尺" in hn:
        return True
    return hn.replace(" ", "") == "面積(平方公尺)"


def _cell_value_for_header(header: Any, ctx: Dict[str, Any]) -> Any:
    """依表頭文字對應內容（與 SAMPLE 內部標題一致）。"""
    if header is None:
        return None
    h_raw = str(header)
    h = h_raw.strip()
    hn = _header_flat(header)

    if h == "編號" or hn == "編號":
        return ctx.get("index_no")
    if h == "地段號" or hn == "地段號":
        return ctx.get("land_section_lot_label", "")
    if h == "地段建號" or hn == "地段建號":
        return ctx.get("building_section_lot_label", "")
    if h == "建物門牌" or hn == "建物門牌":
        return ctx.get("building_doorplate", "")

    # 建物專用面積與用途欄
    if "主建物面積" in hn and "平方公尺" in hn:
        return ctx.get("building_main_area_m2")
    if "附屬建物面積" in hn and "平方公尺" in hn:
        return ctx.get("building_ancillary_m2")
    if "共有部分面積" in hn and "平方公尺" in hn:
        return ctx.get("building_common_m2")
    if "權狀面積" in hn and "平方公尺" in hn:
        return ctx.get("building_cert_m2")
    if hn == "主要用途" or h == "主要用途":
        return ctx.get("building_main_use", "")

    if _header_matches_land_area(header):
        return ctx.get("area_number")
    if h == "所有權人":
        return ctx.get("owner", "")
    if h == "統一編號":
        return ctx.get("id_no", "")
    if h == "地址":
        return ctx.get("address", "")
    if h == "權利範圍":
        return ctx.get("right_range", "")
    if h == "登記原因":
        return ctx.get("own_reg_reason", "")
    if h == "登記日期":
        return ctx.get("own_reg_date", "")
    if h == "查詢時間":
        return ctx.get("query_time", "")
    if _is_remark_header(header):
        return ctx.get("remark", "")
    return None


def _write_ctx_to_row(
    ws,
    row: int,
    ctx: Dict[str, Any],
    header_row: Optional[int] = None,
) -> None:
    """依偵測或指定之表頭列寫入資料列。"""
    hr = header_row if header_row is not None else _detect_header_row(ws)
    # 掃描至足夠欄位；末欄「備註」常在右側，max_column 或舊上限過小會漏寫
    max_scan = min(max((ws.max_column or 0) + 80, 80), 256)
    c = 1
    while c <= max_scan:
        hdr = ws.cell(row=hr, column=c).value
        if hdr is None or str(hdr).strip() == "":
            c += 1
            continue
        ws.cell(row=row, column=c, value=_cell_value_for_header(hdr, ctx))
        c += 1
    # 再次對「備註」欄寫入，避免表頭辨識或掃描邊界與主迴圈不一致
    rmk = ctx.get("remark", "")
    col_rm = _find_remark_column(ws, hr)
    if col_rm is not None:
        ws.cell(row=row, column=col_rm, value="" if rmk is None else str(rmk).strip())


def _create_index_from_sample(dest: Path) -> None:
    """自 SAMPLE 建立 INDEX：土地／建物皆保留表頭列，其下資料刪除；字型設為標楷體。"""
    import openpyxl

    dest.parent.mkdir(parents=True, exist_ok=True)
    sample_path = _resolve_sample_path()
    if not sample_path.exists():
        logger.error("[EXCEL] 無法建立 INDEX：找不到 SAMPLE.xlsx（%s）", sample_path)
        return
    wb = openpyxl.load_workbook(sample_path)
    if "土地" not in wb.sheetnames:
        logger.error("[EXCEL] SAMPLE 缺少「土地」工作表，無法建立 INDEX。")
        wb.close()
        return
    hr = _detect_header_row(wb["土地"])
    _truncate_rows_after_header(wb["土地"], header_last_row=hr)
    _ensure_building_sheet_with_headers(wb)
    if "建物" in wb.sheetnames:
        hr2 = _detect_header_row(wb["建物"])
        _truncate_rows_after_header(wb["建物"], header_last_row=hr2)
    if _FONT_INDEX is not None:
        _apply_kai_font_to_used_range(wb["土地"])
        if "建物" in wb.sheetnames:
            _apply_kai_font_to_used_range(wb["建物"])
    _apply_index_workbook_header_freeze_and_protection(wb)
    wb.save(str(dest))
    wb.close()
    logger.info("[EXCEL] 已自 SAMPLE 建立 INDEX（土地＋建物表頭、標楷體）：%s", dest.name)


def _is_new_index_format(wb) -> bool:
    """INDEX 是否具「土地」「建物」兩分頁，且表頭列含「編號」。"""
    if "土地" not in wb.sheetnames or "建物" not in wb.sheetnames:
        return False
    ws = wb["土地"]
    hr = _detect_header_row(ws)
    return ws.cell(row=hr, column=1).value == "編號"


def _ensure_index_workbook_has_frozen_protected_headers(path: Path) -> None:
    """若檔案存在，重新套用凍結／保護並存檔（供啟動時同步舊檔）。"""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        return
    if not path.is_file():
        return
    import openpyxl

    wb = None
    try:
        wb = openpyxl.load_workbook(path)
        if "土地" not in wb.sheetnames and "建物" not in wb.sheetnames:
            logger.warning(
                "[EXCEL] INDEX 無「土地」「建物」工作表，略過凍結／保護：%s",
                path.name,
            )
            return
        _apply_index_workbook_header_freeze_and_protection(wb)
        wb.save(str(path))
        logger.info("[EXCEL] 已套用 INDEX 表頭凍結與保護：%s", path.name)
    except OSError as e:
        logger.warning(
            "[EXCEL] INDEX 凍結／保護未寫入（檔案是否正被 Excel 或其他程式開啟？）：%s",
            e,
        )
    except Exception as e:
        logger.exception("[EXCEL] INDEX 凍結／保護套用失敗：%s", e)
    finally:
        if wb is not None:
            wb.close()


def _migrate_legacy_index_workbook(path: Path) -> None:
    """將舊版（單一工作表 INDEX、表頭在第 1 列）遷移為 SAMPLE 同款雙工作表格式。"""
    import openpyxl
    import shutil

    wb_old = openpyxl.load_workbook(path, data_only=True)
    try:
        if _LEGACY_INDEX_SHEET not in wb_old.sheetnames:
            return
        ws_old = wb_old[_LEGACY_INDEX_SHEET]
        if ws_old.cell(row=1, column=1).value != "編號":
            return
        rows: List[List[Any]] = []
        for r in range(2, (ws_old.max_row or 1) + 1):
            row = [ws_old.cell(row=r, column=c).value for c in range(1, 7)]
            if row[0] is None or str(row[0]).strip() == "":
                continue
            rows.append(row)
    finally:
        wb_old.close()

    backup = path.with_name("INDEX_舊版備份.xlsx")
    try:
        shutil.copy2(path, backup)
        logger.info("[EXCEL] 舊版 INDEX 已備份為：%s", backup.name)
    except Exception as e:
        logger.warning("[EXCEL] 舊版 INDEX 備份失敗（仍會遷移）：%s", e)

    tmp = path.with_suffix(".tmp.xlsx")
    _create_index_from_sample(tmp)
    wb_new = openpyxl.load_workbook(tmp)
    try:
        for row in rows:
            seq = row[0]
            label = row[3] if len(row) > 3 else None
            qtime = row[5] if len(row) > 5 else None
            try:
                idx = int(float(str(seq).strip()))
            except (TypeError, ValueError):
                continue
            lbl = str(label or "").strip()
            ctx: Dict[str, Any] = {
                "index_no": idx,
                "land_section_lot_label": lbl,
                "building_section_lot_label": "",
                "area_number": None,
                "owner": "",
                "id_no": "",
                "address": "",
                "right_range": "",
                "own_reg_reason": "",
                "own_reg_date": "",
                "query_time": str(qtime).strip() if qtime is not None else "",
                "remark": "",
            }
            ws_tgt = wb_new["土地"]
            _hr = _detect_header_row(ws_tgt)
            tgt_row = _find_first_empty_row_col_a(ws_tgt, start=_hr + 1)
            _write_ctx_to_row(ws_tgt, tgt_row, ctx, header_row=_hr)
        _apply_index_workbook_header_freeze_and_protection(wb_new)
        wb_new.save(str(tmp))
    finally:
        wb_new.close()

    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass
    shutil.move(str(tmp), str(path))
    logger.info("[EXCEL] INDEX 已遷移（土地同 SAMPLE、建物空白、資料僅寫土地）。")


def _ensure_index_workbook(path: Path) -> None:
    """確保 INDEX.xlsx 存在且為 SAMPLE 同款；若為舊版則遷移。"""
    try:
        import openpyxl
    except ImportError:
        return

    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        _create_index_from_sample(path)
        return

    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        if _is_new_index_format(wb):
            return
        need_migrate = _LEGACY_INDEX_SHEET in wb.sheetnames
    finally:
        wb.close()

    if need_migrate:
        _migrate_legacy_index_workbook(path)
    else:
        logger.warning(
            "[EXCEL] INDEX.xlsx 格式無法辨識，將以 SAMPLE 重建（建議先備份）。路徑=%s",
            path,
        )
        bak = path.with_suffix(".bak.xlsx")
        try:
            path.replace(bak)
        except Exception:
            pass
        _create_index_from_sample(path)


def ensure_index_exists() -> None:
    """於主程式啟動時呼叫：INDEX 含土地（表頭同 SAMPLE）與建物（空白）兩分頁。"""
    OUTPUT_BY_SECTION_LOT.mkdir(parents=True, exist_ok=True)
    _ensure_index_workbook(INDEX_XLSX)
    _ensure_index_workbook_has_frozen_protected_headers(INDEX_XLSX)


def ensure_index_exists_at(index_path: Path) -> None:
    """在指定路徑建立（或修復）「INDEX.xlsx 同款格式」。

    目的：讓 Web 可對每個 job 生成獨立輸出檔，而不是覆寫/累積寫死的全域 `INDEX_XLSX`。
    """
    index_path.parent.mkdir(parents=True, exist_ok=True)
    _ensure_index_workbook(index_path)
    _ensure_index_workbook_has_frozen_protected_headers(index_path)


def _find_first_empty_row_col_a(ws, start: int | None = None) -> int:
    """自 start 列起找 A 欄第一列空白（供 INDEX 累積新增）。"""
    if start is None:
        start = _detect_header_row(ws) + 1
    r = start
    while r < 100000:
        v = ws.cell(row=r, column=1).value
        if v is None or str(v).strip() == "":
            return r
        r += 1
    return r


def _read_max_index_number(path: Path, sheet_name: str = "土地") -> int:
    """讀取指定工作表 A 欄（表頭下一列起）之最大編號。"""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        mx = 0
        if sheet_name not in wb.sheetnames:
            return 0
        ws = wb[sheet_name]
        hr = _detect_header_row(ws)
        for r in range(hr + 1, (ws.max_row or hr) + 1):
            v = ws.cell(row=r, column=1).value
            if v is None or str(v).strip() == "":
                continue
            try:
                mx = max(mx, int(float(str(v).strip())))
            except ValueError:
                continue
        return mx
    finally:
        wb.close()


def _append_index_row(path: Path, sheet_name: str, ctx: Dict[str, Any]) -> None:
    """於 INDEX 指定工作表附加一列，並將該列字型設為標楷體。"""
    import openpyxl

    wb = openpyxl.load_workbook(path)
    try:
        # 切勿於此呼叫 _ensure_building_sheet_with_headers：其會截斷表頭以下列，清空已累積 INDEX
        if sheet_name not in wb.sheetnames:
            logger.error("[EXCEL] INDEX 缺少工作表：%s", sheet_name)
            return
        ws = wb[sheet_name]
        hr = _detect_header_row(ws)
        r = _find_first_empty_row_col_a(ws, start=hr + 1)
        _write_ctx_to_row(ws, r, ctx, header_row=hr)
        if _FONT_INDEX is not None:
            mc_font = min(max(ws.max_column or 0, 48), 256)
            for cc in range(1, mc_font + 1):
                cell = ws.cell(row=r, column=cc)
                if cell.value is not None and str(cell.value).strip() != "":
                    cell.font = _FONT_INDEX
        _apply_index_workbook_header_freeze_and_protection(wb)
        wb.save(str(path))
    finally:
        wb.close()


def _append_index_row_to_open_workbook(
    wb,
    *,
    sheet_name: str,
    ctx: Dict[str, Any],
) -> None:
    """在已開啟的工作簿中追加一列（不立即 save，供批次寫入使用）。"""
    if sheet_name not in wb.sheetnames:
        logger.error("[EXCEL] INDEX 缺少工作表：%s", sheet_name)
        return
    ws = wb[sheet_name]
    hr = _detect_header_row(ws)
    r = _find_first_empty_row_col_a(ws, start=hr + 1)
    _write_ctx_to_row(ws, r, ctx, header_row=hr)
    if _FONT_INDEX is not None:
        mc_font = min(max(ws.max_column or 0, 48), 256)
        for cc in range(1, mc_font + 1):
            cell = ws.cell(row=r, column=cc)
            if cell.value is not None and str(cell.value).strip() != "":
                cell.font = _FONT_INDEX


def _make_row_context(
    *,
    doc_kind: str,
    index_no: int,
    section: str,
    lot: str,
    address: str,
    excel_data_mark: Dict[str, Any],
    excel_data_ownership: Dict[str, Any],
    remark: str = "",
) -> Dict[str, Any]:
    """組出與 SAMPLE 橫向欄位對應之一列內容（供個案檔與 INDEX 共用）；備註由呼叫端傳入（預設空白）。"""
    mk = excel_data_mark or {}
    ow = excel_data_ownership or {}
    land_lbl = mk.get("地段地號", "")
    if not land_lbl and (section or lot):
        sfx = "建號" if str(doc_kind) == "building" else "地號"
        land_lbl = f"{section} {lot} {sfx}".strip()
    # 建物表未定義時，建號亦寫入「土地」表之「地段號」欄，故不重複使用 building 專用欄
    bld_lbl = ""

    area_num = _parse_area_sqm_number(mk.get("面積"))
    q_own = ow.get("查詢時間", "") or ""
    q_mk = mk.get("查詢時間", "") or ""
    query_time = q_own or q_mk

    return {
        "index_no": int(index_no),
        "land_section_lot_label": land_lbl,
        "building_section_lot_label": bld_lbl,
        "area_number": area_num,
        "owner": ow.get("所有權人", ""),
        "id_no": ow.get("統一編號", ""),
        "address": address or "",
        "right_range": ow.get("權利範圍", ""),
        "own_reg_reason": ow.get("登記原因", ""),
        "own_reg_date": ow.get("登記日期", ""),
        "query_time": query_time,
        "remark": str(remark or "").strip(),
    }


def _make_building_row_context(
    *,
    index_no: int,
    excel_data_building: Dict[str, Any],
    excel_data_ownership: Dict[str, Any],
    address: str,
    remark: str = "",
) -> Dict[str, Any]:
    """建物 INDEX／輸出列：門牌、各面積、主要用途＋所有權部欄位。"""
    bd = excel_data_building or {}
    ow = excel_data_ownership or {}
    q_own = ow.get("查詢時間", "") or ""
    q_bd = bd.get("查詢時間", "") or ""
    query_time = q_own or q_bd
    bsec = str(bd.get("地段建號") or "").strip()
    return {
        "index_no": int(index_no),
        "land_section_lot_label": "",
        "building_section_lot_label": bsec,
        "area_number": None,
        "building_doorplate": bd.get("建物門牌", ""),
        "building_main_area_m2": bd.get("主建物面積"),
        "building_ancillary_m2": bd.get("附屬建物面積"),
        "building_common_m2": bd.get("共有部分面積"),
        "building_cert_m2": bd.get("權狀面積"),
        "building_main_use": bd.get("主要用途", ""),
        "owner": ow.get("所有權人", ""),
        "id_no": ow.get("統一編號", ""),
        "address": address or "",
        "right_range": ow.get("權利範圍", ""),
        "own_reg_reason": ow.get("登記原因", ""),
        "own_reg_date": ow.get("登記日期", ""),
        "query_time": query_time,
        "remark": str(remark or "").strip(),
    }


def prune_output_excel_extras(output_dir: Optional[Path] = None) -> None:
    """刪除輸出目錄內除 SAMPLE.xlsx、INDEX.xlsx 以外之 Excel（個案檔不再使用）。"""
    root = output_dir or OUTPUT_BY_SECTION_LOT
    if not root.is_dir():
        return
    keep = {SAMPLE_XLSX.resolve().name, INDEX_XLSX.resolve().name}
    for f in root.glob("*.xlsx"):
        if f.name.startswith("~$"):
            continue
        if f.name in keep:
            continue
        try:
            f.unlink()
            logger.info("[EXCEL] 已移除多餘 Excel: %s", f.name)
        except OSError as e:
            logger.warning("[EXCEL] 無法刪除 %s: %s", f.name, e)


def _coerce_optional_float(val: Any) -> Any:
    """將表單字串轉為 float；空值回傳 None。"""
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    try:
        return float(str(val).replace(",", "").strip())
    except (TypeError, ValueError):
        return val


def save_user_confirmed_row_to_index(
    *,
    doc_kind: str,
    section: str,
    lot: str,
    excel_data_mark: Dict[str, Any],
    excel_data_ownership: Dict[str, Any],
    excel_data_building: Optional[Dict[str, Any]] = None,
    remark: str = "",
) -> int:
    """使用者於前端確認後寫入 INDEX；回傳本列編號。remark 對應試算表「備註」欄。"""
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        raise RuntimeError("openpyxl 未安裝，無法寫入 INDEX")

    bd_in = excel_data_building or {}
    bd: Dict[str, Any] = dict(bd_in)
    if bd:
        for k in ("主建物面積", "附屬建物面積", "共有部分面積", "權狀面積"):
            if k in bd:
                bd[k] = _coerce_optional_float(bd.get(k))

    mk = dict(excel_data_mark or {})
    ow = dict(excel_data_ownership or {})
    address = str(ow.get("地址", "") or "").strip()

    with _INDEX_LOCK:
        _ensure_index_workbook(INDEX_XLSX)
        if str(doc_kind) == "building":
            sheet_nm = "建物"
            next_no = _read_max_index_number(INDEX_XLSX, "建物") + 1
            ctx = _make_building_row_context(
                index_no=next_no,
                excel_data_building=bd,
                excel_data_ownership=ow,
                address=address,
                remark=remark,
            )
        else:
            sheet_nm = "土地"
            next_no = _read_max_index_number(INDEX_XLSX, "土地") + 1
            ctx = _make_row_context(
                doc_kind=str(doc_kind),
                index_no=next_no,
                section=section,
                lot=lot,
                address=address,
                excel_data_mark=mk,
                excel_data_ownership=ow,
                remark=remark,
            )
        _append_index_row(INDEX_XLSX, sheet_nm, ctx)
    logger.info(
        "[EXCEL] 使用者確認寫入 INDEX[%s] 編號 %s",
        sheet_nm,
        next_no,
    )
    return int(next_no)


def write_result_by_section_lot(result: Dict[str, Any]) -> Path:
    """僅將一筆結果附加至 INDEX.xlsx（土地／建物分頁）；不產生其他 Excel。"""
    OUTPUT_BY_SECTION_LOT.mkdir(parents=True, exist_ok=True)
    section = result.get("地段", "")
    lot = result.get("地號", "")
    doc_kind = result.get("doc_kind") or "land"
    base = safe_filename(section, lot)

    excel_mark = result.get("excel_data_mark") or {}
    excel_own = result.get("excel_data_ownership") or {}
    address = excel_own.get("地址", result.get("address", ""))

    excel_bld = result.get("excel_data_building") or {}

    try:
        return write_result_by_section_lot_to(
            result=result,
            index_path=INDEX_XLSX,
            backup_dir=OUTPUT_BY_SECTION_LOT,
        )
    except Exception:
        # 保險：舊版 API 不要讓例外一路泡出（避免 Web 端整批崩潰）。
        # 實際例外細節會由 write_result_by_section_lot_to 內部 log。
        return INDEX_XLSX


def write_result_by_section_lot_to(
    *,
    result: Dict[str, Any],
    index_path: Path,
    backup_dir: Optional[Path] = None,
) -> Path:
    """將一筆結果附加至指定 INDEX-like Excel 檔案。

    - Web job：每個 job 會有自己的 `index_path`，用於產生可下載的獨立 EXCEL。
    - 備援文字檔會寫入 `backup_dir`（預設為 index_path.parent）。
    """
    section = result.get("地段", "")
    lot = result.get("地號", "")
    doc_kind = result.get("doc_kind") or "land"
    base = safe_filename(section, lot)

    excel_mark = result.get("excel_data_mark") or {}
    excel_own = result.get("excel_data_ownership") or {}
    address = excel_own.get("地址", result.get("address", ""))
    excel_bld = result.get("excel_data_building") or {}

    backup_dir = backup_dir or index_path.parent
    backup_dir.mkdir(parents=True, exist_ok=True)

    try:
        import openpyxl  # noqa: F401  # 確認已安裝
    except ImportError:
        logger.error(
            "[EXCEL] openpyxl 未安裝，無法寫入 INDEX-like 檔。請執行：pip install openpyxl"
        )
        openpyxl = None  # type: ignore

    if openpyxl is not None:
        with _INDEX_LOCK:
            ensure_index_exists_at(index_path)
            if str(doc_kind) == "building":
                sheet_nm = "建物"
                next_no = _read_max_index_number(index_path, "建物") + 1
                ctx_index = _make_building_row_context(
                    index_no=next_no,
                    excel_data_building=excel_bld,
                    excel_data_ownership=excel_own,
                    address=address or "",
                )
            else:
                sheet_nm = "土地"
                next_no = _read_max_index_number(index_path, "土地") + 1
                ctx_index = _make_row_context(
                    doc_kind=str(doc_kind),
                    index_no=next_no,
                    section=section,
                    lot=lot,
                    address=address or "",
                    excel_data_mark=excel_mark,
                    excel_data_ownership=excel_own,
                )

            try:
                _append_index_row(index_path, sheet_nm, ctx_index)
                return index_path
            except Exception as e:
                logger.error(
                    "[EXCEL] 寫入 INDEX-like 失敗: %s 原因=%s",
                    index_path,
                    e,
                )

    # INDEX-like 失敗時備援文字檔
    out_path_txt = backup_dir / (base + ".txt")
    try:
        with open(out_path_txt, "w", encoding="utf-8") as f:
            f.write(result.get("content", ""))
        return out_path_txt
    except Exception:
        return out_path_txt


def write_results_batch_by_section_lot(
    results: List[Dict[str, Any]],
    *,
    index_path: Path = INDEX_XLSX,
) -> Path:
    """批次寫入多筆結果至 INDEX，僅於批次尾端一次 save + 套用保護。

    設計目的：
    - 降低每筆都開關檔與重套保護造成的 I/O 成本
    - 供 main.py 的單一 Writer 執行緒呼叫，達成「邊解譯邊寫」
    """
    if not results:
        return index_path

    try:
        import openpyxl
    except ImportError as ex:
        raise RuntimeError("openpyxl 未安裝，無法批次寫入 INDEX") from ex

    with _INDEX_LOCK:
        ensure_index_exists_at(index_path)
        wb = openpyxl.load_workbook(index_path)
        try:
            if "土地" not in wb.sheetnames or "建物" not in wb.sheetnames:
                raise RuntimeError("INDEX 缺少必要工作表（土地/建物）")

            ws_land = wb["土地"]
            ws_build = wb["建物"]
            hr_land = _detect_header_row(ws_land)
            hr_build = _detect_header_row(ws_build)

            # 批次內連號：先抓現況最大編號，後續逐筆遞增
            max_land = 0
            for r in range(hr_land + 1, (ws_land.max_row or hr_land) + 1):
                v = ws_land.cell(row=r, column=1).value
                if v is None or str(v).strip() == "":
                    continue
                try:
                    max_land = max(max_land, int(float(str(v).strip())))
                except ValueError:
                    continue

            max_build = 0
            for r in range(hr_build + 1, (ws_build.max_row or hr_build) + 1):
                v = ws_build.cell(row=r, column=1).value
                if v is None or str(v).strip() == "":
                    continue
                try:
                    max_build = max(max_build, int(float(str(v).strip())))
                except ValueError:
                    continue

            for result in results:
                section = result.get("地段", "")
                lot = result.get("地號", "")
                doc_kind = result.get("doc_kind") or "land"
                excel_mark = result.get("excel_data_mark") or {}
                excel_own = result.get("excel_data_ownership") or {}
                address = excel_own.get("地址", result.get("address", ""))
                excel_bld = result.get("excel_data_building") or {}

                if str(doc_kind) == "building":
                    max_build += 1
                    ctx = _make_building_row_context(
                        index_no=max_build,
                        excel_data_building=excel_bld,
                        excel_data_ownership=excel_own,
                        address=address or "",
                    )
                    _append_index_row_to_open_workbook(
                        wb, sheet_name="建物", ctx=ctx
                    )
                else:
                    max_land += 1
                    ctx = _make_row_context(
                        doc_kind=str(doc_kind),
                        index_no=max_land,
                        section=section,
                        lot=lot,
                        address=address or "",
                        excel_data_mark=excel_mark,
                        excel_data_ownership=excel_own,
                    )
                    _append_index_row_to_open_workbook(
                        wb, sheet_name="土地", ctx=ctx
                    )

            # 僅在批次尾端套用一次保護與存檔
            _apply_index_workbook_header_freeze_and_protection(wb)
            wb.save(str(index_path))
            logger.info("[EXCEL] 批次寫入完成：%d 筆 -> %s", len(results), index_path.name)
        finally:
            wb.close()
    return index_path
